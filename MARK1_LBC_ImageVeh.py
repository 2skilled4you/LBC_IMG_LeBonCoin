from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
import time
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import urllib
import datetime
import linecache
import sys
from subprocess import Popen, PIPE, STDOUT, check_output

#strFullURLName = "C:/Users/petitbrandao/Documents/Python Scripts/03. Ebay Allianz/Reverse Tarif Allianz.xlsx"
#wbEbay = load_workbook(filename=strFullURLName)
#
#print(wbAPRIL.get_sheet_names())
#
#wsEbay = wbEbay.get_sheet_by_name('Data')
#wsParametres = wbAPRIL.get_sheet_by_name('Paramètres')
#print(wsGroupama['A1'].value)
#print(wsGroupama['A' + str(1)].value)
############################################################
######## FONCTIONS UTILES AU DEROULEMENT DU SCRIPT #########
############################################################

# 1) Fonction qui donne des informations détaillées sur l'erreur produite 
#   par le script
def PrintException():
    exc_type, exc_obj, tb = sys.exc_info()
    f = tb.tb_frame
    lineno = tb.tb_lineno
    filename = f.f_code.co_filename
    linecache.checkcache(filename)
    line = linecache.getline(filename, lineno, f.f_globals)
    print('EXCEPTION IN ({}, LINE {} "{}"): {}'.format(filename, lineno, line.strip(), exc_obj))

# 2) Fonction qui permettent de nettoyer le cache du navigateur
def get_clear_browsing_button(driver):
    """Find the "CLEAR BROWSING BUTTON" on the Chrome settings page."""
    return driver.find_element_by_css_selector('* /deep/ #clearBrowsingDataConfirm')


def clear_cache(driver, timeout=60):
    """Clear the cookies and cache for the ChromeDriver instance."""
    # navigate to the settings page
    driver.get('chrome://settings/clearBrowserData')

    # wait for the button to appear
    wait = WebDriverWait(driver, timeout)
    wait.until(get_clear_browsing_button)

    # click the button to clear the cache
    get_clear_browsing_button(driver).click()

    # wait for the button to be gone before returning
    wait.until_not(get_clear_browsing_button)
   
# 3) Fonction qui permet de tester la présence d'un élement dans la page
def check_exists_by_xpath(xpath):
    try:
        driver.find_element_by_xpath(xpath)
    except NoSuchElementException:
        return False
    return True

py_process = Popen([r'C:\Users\Bernard\Documents\Python Scripts\60. Outils\clearcachePACIFICA.bat'], stdout=PIPE, stderr=STDOUT)

time.sleep(5)
#subprocess.call([r'C:\Users\Bernard\Documents\Python Scripts\80. Git\LauchChromeExperimental.bat'])
#            OpenChromeExp = Popen('cmd')
#            Popen('"C:\Program Files\Google\Chrome\Application\chrome.exe" --remote-debugging-port=9223 --user-data-dir="C:\selenium\AutomationProfile" "http://127.0.0.1:9223"')
py_process = Popen([r'C:\Users\Bernard\Documents\Python Scripts\60. Outils\LaunchChromeExperimentalPACIFICA.bat'], stdout=PIPE, stderr=STDOUT)

#driver = webdriver.Firefox()
chrome_driver = "C:\\Users\\Public\\Documents\\Python Scripts\\chromedriver.exe"
chrome_options = Options()
chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9225")

time.sleep(10)
            
driver = webdriver.Chrome(chrome_driver, chrome_options=chrome_options)

#driver = webdriver.Chrome("C:\\Users\\Bernard\\chromedriver.exe")

driver.get("https://www.leboncoin.fr/voitures/offres/")

time.sleep(5)

for eachButton in driver.find_elements_by_css_selector("button"):
    print(eachButton.get_attribute("innerText"))
    if eachButton.get_attribute("innerText") == "J’ai compris":
        eachButton.click()
        
#inpPRO = driver.find_element_by_xpath("//div[@data-qa-id='pro_filter']/div/label/span[2]")
#inpPRO.click()

time.sleep(5)

strImgPath = "C:/Users/Public/Documents/Drive/04. Reconnaissance d'image - Immatriculations/02. Collecte/02. ImageVeh/"

for i in range(0, 1):
    inpPage = driver.find_element_by_id("container")
    listeAnnonce = inpPage.find_elements_by_xpath(".//ul/li")
    for eachAnnonce in listeAnnonce:
        if "€" in eachAnnonce.text:
            driver_temp = webdriver.Chrome("C:\\Users\\Bernard\\chromedriver.exe")
            driver_temp.get(eachAnnonce.find_element_by_xpath(".//a").get_attribute("href"))
            time.sleep(5)
            # get the image source
            inpPageDeep = driver_temp.find_element_by_id("container")
            listeImg = inpPageDeep.find_elements_by_xpath(".//img")
            for eachImg in listeImg:
                if "image-galerie" in eachImg.get_attribute("alt"):
                    src = eachImg.get_attribute('src')
                    strTime = str(datetime.datetime.now()).replace(" ","-").replace(":", "-").replace(".","-")
                    urllib.request.urlretrieve(src, strImgPath + strTime +".png")
                    time.sleep(1)
            driver_temp.quit()
            time.sleep(5)      
    time.sleep(60)
    driver.refresh()
    time.sleep(10)
#labelAnnonce = inpAnnonce.get_attribute("innerText")
#print(labelAnnonce)



